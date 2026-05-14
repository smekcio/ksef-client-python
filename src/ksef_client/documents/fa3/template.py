from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any

from .models import FA3ImportResult, FA3ValidationIssue

SHEET_INVOICES = "Faktury"
SHEET_HELP = "Pomoc"
SHEET_DICTIONARIES = "Słowniki"
HEADER_ROW = 2
DATA_START_ROW = 3


@dataclass(frozen=True)
class ColumnSpec:
    name: str
    section: str
    width: int
    hidden: bool = False
    number_format: str | None = None
    dictionary: str | None = None


COLUMNS: tuple[ColumnSpec, ...] = (
    ColumnSpec("Numer faktury", "Faktura", 22),
    ColumnSpec("Typ faktury", "Faktura", 26, dictionary="typy"),
    ColumnSpec("Data wystawienia", "Faktura", 16, number_format="yyyy-mm-dd"),
    ColumnSpec("Waluta", "Faktura", 12, dictionary="waluty"),
    ColumnSpec("Miejsce wystawienia", "Faktura", 22),
    ColumnSpec("NIP sprzedawcy", "Sprzedawca", 18),
    ColumnSpec("Nazwa sprzedawcy", "Sprzedawca", 28),
    ColumnSpec("Adres sprzedawcy", "Sprzedawca", 36),
    ColumnSpec("Kraj sprzedawcy", "Sprzedawca", 14, dictionary="kraje"),
    ColumnSpec("NIP nabywcy", "Nabywca", 18),
    ColumnSpec("Nazwa nabywcy", "Nabywca", 28),
    ColumnSpec("Adres nabywcy", "Nabywca", 36),
    ColumnSpec("Kraj nabywcy", "Nabywca", 14, dictionary="kraje"),
    ColumnSpec("Opis pozycji", "Pozycja", 36),
    ColumnSpec("Ilość", "Pozycja", 12, number_format="0.0000"),
    ColumnSpec("Jm", "Pozycja", 10),
    ColumnSpec("Cena netto", "Pozycja", 14, number_format="#,##0.00"),
    ColumnSpec("VAT", "VAT i kwoty", 12, dictionary="vat"),
    ColumnSpec("Wartość netto", "VAT i kwoty", 15, number_format="#,##0.00"),
    ColumnSpec("Kwota VAT", "VAT i kwoty", 14, number_format="#,##0.00"),
    ColumnSpec("Wartość brutto", "VAT i kwoty", 15, number_format="#,##0.00"),
    ColumnSpec("Termin płatności", "Płatność", 16, number_format="yyyy-mm-dd"),
    ColumnSpec("Forma płatności", "Płatność", 20, dictionary="platnosci"),
    ColumnSpec("Przyczyna korekty", "Korekta/Zaliczka/Rozliczenie", 30, hidden=True),
    ColumnSpec("Numer faktury korygowanej", "Korekta/Zaliczka/Rozliczenie", 26, hidden=True),
    ColumnSpec(
        "Data faktury korygowanej",
        "Korekta/Zaliczka/Rozliczenie",
        24,
        hidden=True,
        number_format="yyyy-mm-dd",
    ),
    ColumnSpec("Numer KSeF faktury korygowanej", "Korekta/Zaliczka/Rozliczenie", 30, hidden=True),
    ColumnSpec("Numer faktury zaliczkowej", "Korekta/Zaliczka/Rozliczenie", 26, hidden=True),
    ColumnSpec("Numer KSeF faktury zaliczkowej", "Korekta/Zaliczka/Rozliczenie", 30, hidden=True),
    ColumnSpec(
        "Kwota rozliczenia",
        "Korekta/Zaliczka/Rozliczenie",
        18,
        hidden=True,
        number_format="#,##0.00",
    ),
)

REQUIRED_HEADERS = tuple(column.name for column in COLUMNS)

DICTIONARIES = {
    "typy": [
        "podstawowa",
        "uproszczona",
        "korygująca",
        "zaliczkowa",
        "rozliczeniowa",
        "korekta zaliczki",
        "korekta rozliczenia",
    ],
    "vat": ["23", "8", "5", "0", "zw"],
    "waluty": ["PLN", "EUR", "USD", "GBP", "CHF", "CZK", "SEK", "NOK", "DKK"],
    "kraje": ["PL", "DE", "CZ", "SK", "FR", "IT", "ES", "NL", "GB", "US"],
    "platnosci": ["przelew", "gotówka", "karta", "kompensata", "inny"],
}


def _require_openpyxl() -> Any:
    try:
        import openpyxl
    except ImportError as exc:  # pragma: no cover - covered by optional dependency packaging.
        raise RuntimeError(
            'Obsługa XLSX wymaga openpyxl. Zainstaluj: pip install "ksef-client[fa3-xlsx]".'
        ) from exc
    return openpyxl


class FA3Template:
    @staticmethod
    def create_xlsx(path: str | Path, sample: bool = True) -> None:
        openpyxl = _require_openpyxl()
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.datavalidation import DataValidation

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = SHEET_INVOICES
        help_ws = wb.create_sheet(SHEET_HELP)
        dict_ws = wb.create_sheet(SHEET_DICTIONARIES)

        section_fill = PatternFill("solid", fgColor="1F4E78")
        header_fill = PatternFill("solid", fgColor="D9EAF7")
        header_font = Font(bold=True, color="000000")
        section_font = Font(bold=True, color="FFFFFF")

        for index, column in enumerate(COLUMNS, start=1):
            letter = get_column_letter(index)
            ws.cell(row=1, column=index, value=column.section)
            ws.cell(row=2, column=index, value=column.name)
            ws.column_dimensions[letter].width = column.width
            ws.column_dimensions[letter].hidden = column.hidden
            ws.cell(row=1, column=index).fill = section_fill
            ws.cell(row=1, column=index).font = section_font
            ws.cell(row=2, column=index).fill = header_fill
            ws.cell(row=2, column=index).font = header_font
            ws.cell(row=2, column=index).alignment = Alignment(wrap_text=True)
            if column.number_format:
                for row in range(DATA_START_ROW, 1001):
                    ws.cell(row=row, column=index).number_format = column.number_format

        ws.freeze_panes = "A3"
        ws.auto_filter.ref = f"A2:{get_column_letter(len(COLUMNS))}2"

        _populate_dictionaries(dict_ws)
        _add_data_validations(ws, dict_ws, DataValidation)
        _populate_help(help_ws)
        if sample:
            _populate_sample(ws)

        dict_ws.sheet_state = "hidden"
        wb.save(path)


def _populate_dictionaries(ws: Any) -> None:
    for col_index, (name, values) in enumerate(DICTIONARIES.items(), start=1):
        ws.cell(row=1, column=col_index, value=name)
        for row_index, value in enumerate(values, start=2):
            ws.cell(row=row_index, column=col_index, value=value)


def _add_data_validations(ws: Any, dict_ws: Any, data_validation_cls: Any) -> None:
    from openpyxl.utils import get_column_letter

    dictionary_columns = {name: index for index, name in enumerate(DICTIONARIES, start=1)}
    for index, column in enumerate(COLUMNS, start=1):
        if not column.dictionary:
            continue
        source_col = dictionary_columns[column.dictionary]
        source_letter = get_column_letter(source_col)
        values_count = len(DICTIONARIES[column.dictionary]) + 1
        formula = f"'{SHEET_DICTIONARIES}'!${source_letter}$2:${source_letter}${values_count}"
        validation = data_validation_cls(type="list", formula1=formula, allow_blank=True)
        validation.errorTitle = "Nieprawidłowa wartość"
        validation.error = "Wybierz wartość z listy."
        ws.add_data_validation(validation)
        validation.add(f"{get_column_letter(index)}{DATA_START_ROW}:{get_column_letter(index)}1000")


def _populate_help(ws: Any) -> None:
    rows = [
        ("Zasada", "Jeden wiersz to jedna pozycja faktury."),
        (
            "Wiele pozycji",
            "Powtórz numer faktury; dane nagłówka możesz wpisać tylko w pierwszym wierszu.",
        ),
        ("Kwoty", "Jeśli zostawisz kwoty puste, SDK wyliczy netto, VAT i brutto."),
        ("Override", "Jeśli wpiszesz własne kwoty różne od wyliczeń, import zgłosi ostrzeżenie."),
        (
            "Zaawansowane",
            "Kolumny korekt, zaliczek i rozliczeń są ukryte; można je odkryć w Excelu.",
        ),
    ]
    ws.append(("Temat", "Opis"))
    for row in rows:
        ws.append(row)
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 100


def _populate_sample(ws: Any) -> None:
    sample = {
        "Numer faktury": "FV/1/2026",
        "Typ faktury": "podstawowa",
        "Data wystawienia": "2026-01-15",
        "Waluta": "PLN",
        "Miejsce wystawienia": "Warszawa",
        "NIP sprzedawcy": "1234567890",
        "Nazwa sprzedawcy": "Przykładowy Sprzedawca sp. z o.o.",
        "Adres sprzedawcy": "ul. Prosta 1, 00-001 Warszawa",
        "Kraj sprzedawcy": "PL",
        "NIP nabywcy": "1111111111",
        "Nazwa nabywcy": "Przykładowy Nabywca sp. z o.o.",
        "Adres nabywcy": "ul. Testowa 2, 00-002 Warszawa",
        "Kraj nabywcy": "PL",
        "Opis pozycji": "Usługa konsultingowa",
        "Ilość": 1,
        "Jm": "szt",
        "Cena netto": 100,
        "VAT": "23",
        "Termin płatności": "2026-01-29",
        "Forma płatności": "przelew",
    }
    for column_index, column in enumerate(COLUMNS, start=1):
        ws.cell(row=DATA_START_ROW, column=column_index, value=sample.get(column.name))


def create_error_report_xlsx(result: FA3ImportResult, path: str | Path) -> None:
    openpyxl = _require_openpyxl()
    from openpyxl.styles import Font, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Raport błędów"
    headers = list(result.headers or REQUIRED_HEADERS)
    report_headers = headers + ["Status", "Błędy", "Ostrzeżenia"]
    ws.append(["Raport importu FA(3)"])
    ws.append(report_headers)

    issues_by_row = _issues_by_row(result)
    source_rows = result.source_rows or []
    for source in source_rows:
        row_number = int(source.get("__row_number__", 0))
        row_errors, row_warnings = issues_by_row.get(row_number, ([], []))
        status = "Błąd" if row_errors else "OK z ostrzeżeniami" if row_warnings else "OK"
        ws.append(
            [source.get(header) for header in headers]
            + [
                status,
                "\n".join(_issue_message(issue) for issue in row_errors),
                "\n".join(_issue_message(issue) for issue in row_warnings),
            ]
        )

    red_fill = PatternFill("solid", fgColor="F8CBAD")
    yellow_fill = PatternFill("solid", fgColor="FFF2CC")
    ws.cell(row=1, column=1).font = Font(bold=True)
    for cell in ws[2]:
        cell.font = Font(bold=True)
    header_to_index = {header: index + 1 for index, header in enumerate(headers)}
    for row_number, (row_errors, row_warnings) in issues_by_row.items():
        excel_row = _source_row_to_report_row(source_rows, row_number)
        if excel_row is None:
            continue
        for issue in row_errors:
            if issue.column in header_to_index:
                ws.cell(row=excel_row, column=header_to_index[issue.column]).fill = red_fill
        for issue in row_warnings:
            if issue.column in header_to_index:
                ws.cell(row=excel_row, column=header_to_index[issue.column]).fill = yellow_fill
    wb.save(path)


def _issues_by_row(
    result: FA3ImportResult,
) -> dict[int, tuple[list[FA3ValidationIssue], list[FA3ValidationIssue]]]:
    rows: dict[int, tuple[list[FA3ValidationIssue], list[FA3ValidationIssue]]] = {}
    for invalid in result.invalid_rows:
        rows.setdefault(invalid.row_number, ([], []))[0].extend(invalid.errors)
        rows.setdefault(invalid.row_number, ([], []))[1].extend(invalid.warnings)
    for warning in result.warnings:
        if warning.row_number is not None:
            rows.setdefault(warning.row_number, ([], []))[1].append(warning)
    return rows


def _source_row_to_report_row(source_rows: list[dict[str, Any]], row_number: int) -> int | None:
    for offset, row in enumerate(source_rows, start=3):
        if row.get("__row_number__") == row_number:
            return offset
    return None


def _issue_message(issue: FA3ValidationIssue) -> str:
    prefix = f"{issue.cell}: " if issue.cell else ""
    return f"{prefix}{issue.message}"
