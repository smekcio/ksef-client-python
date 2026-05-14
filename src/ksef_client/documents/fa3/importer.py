from __future__ import annotations

import json
from dataclasses import replace
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from typing import Any

from .models import (
    FA3BatchDraft,
    FA3Draft,
    FA3ImportResult,
    FA3InvalidRow,
    FA3InvoiceBuilder,
    FA3InvoiceKind,
    FA3Line,
    FA3Party,
    FA3ValidationIssue,
    ImportMode,
    decimal_from_value,
    parse_vat_rate,
)
from .template import (
    DATA_START_ROW,
    HEADER_ROW,
    REQUIRED_HEADERS,
    SHEET_INVOICES,
    _require_openpyxl,
)


class FA3ImportError(ValueError):
    def __init__(self, issue: FA3ValidationIssue) -> None:
        super().__init__(issue.message)
        self.issue = issue


class FA3Importer:
    @staticmethod
    def from_xlsx(
        path: str | Path,
        mode: ImportMode = ImportMode.PARTIAL_WITH_REPORT,
    ) -> FA3ImportResult:
        openpyxl = _require_openpyxl()
        workbook = openpyxl.load_workbook(path, data_only=True)
        if SHEET_INVOICES not in workbook.sheetnames:
            issue = FA3ValidationIssue(f"Brakuje arkusza '{SHEET_INVOICES}'.")
            return _finish_with_structural_error(issue, mode)
        ws = workbook[SHEET_INVOICES]
        headers = [
            ws.cell(row=HEADER_ROW, column=index).value
            for index in range(1, len(REQUIRED_HEADERS) + 1)
        ]
        header_names = [str(header or "").strip() for header in headers]
        missing = [header for header in REQUIRED_HEADERS if header not in header_names]
        if missing:
            issue = FA3ValidationIssue(
                "Arkusz nie jest zgodny z szablonem. Brak kolumn: " + ", ".join(missing) + "."
            )
            return _finish_with_structural_error(issue, mode, headers=header_names)
        if header_names != list(REQUIRED_HEADERS):
            issue = FA3ValidationIssue(
                "Arkusz nie jest zgodny z szablonem. Kolumny muszą być w kolejności "
                "z oficjalnego szablonu."
            )
            return _finish_with_structural_error(issue, mode, headers=header_names)

        rows = _read_rows(ws, header_names)
        result = _drafts_from_rows(rows, header_names, mode)
        if mode is ImportMode.VALIDATE_ONLY:
            result.valid_drafts = []
        return result

    @staticmethod
    def from_json(
        path: str | Path,
        mode: ImportMode = ImportMode.PARTIAL_WITH_REPORT,
    ) -> FA3ImportResult:
        payload = json.loads(Path(path).read_text(encoding="utf-8"))
        try:
            batch = FA3BatchDraft.from_json(payload)
        except (TypeError, ValueError) as exc:
            issue = FA3ValidationIssue(f"JSON draft jest nieprawidłowy: {exc}.")
            return _finish_with_structural_error(issue, mode)

        valid: list[FA3Draft] = []
        invalid: list[FA3InvalidRow] = []
        errors: list[FA3ValidationIssue] = []
        warnings: list[FA3ValidationIssue] = []
        for index, draft in enumerate(batch.drafts, start=1):
            draft_errors, draft_warnings = draft.validate()
            warnings.extend(draft_warnings)
            if draft_errors:
                errors.extend(draft_errors)
                invalid.append(
                    FA3InvalidRow(
                        row_number=index,
                        invoice_number=draft.invoice_number,
                        row_data=draft.to_dict(),
                        errors=tuple(draft_errors),
                        warnings=tuple(draft_warnings),
                    )
                )
                if mode is ImportMode.FAIL_FAST:
                    raise FA3ImportError(draft_errors[0])
            elif mode is not ImportMode.VALIDATE_ONLY:
                valid.append(draft)
        return FA3ImportResult(
            valid_drafts=valid,
            invalid_rows=invalid,
            errors=errors,
            warnings=warnings,
        )


def _finish_with_structural_error(
    issue: FA3ValidationIssue,
    mode: ImportMode,
    *,
    headers: list[str] | None = None,
) -> FA3ImportResult:
    if mode is ImportMode.FAIL_FAST:
        raise FA3ImportError(issue)
    return FA3ImportResult(errors=[issue], headers=headers or [])


def _read_rows(ws: Any, headers: list[str]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for row_number in range(DATA_START_ROW, ws.max_row + 1):
        values = {
            header: ws.cell(row=row_number, column=index + 1).value
            for index, header in enumerate(headers)
        }
        if not any(value not in (None, "") for value in values.values()):
            continue
        values["__row_number__"] = row_number
        rows.append(values)
    return rows


def _drafts_from_rows(
    rows: list[dict[str, Any]],
    headers: list[str],
    mode: ImportMode,
) -> FA3ImportResult:
    valid: list[FA3Draft] = []
    invalid_rows: list[FA3InvalidRow] = []
    errors: list[FA3ValidationIssue] = []
    warnings: list[FA3ValidationIssue] = []
    source_rows: list[dict[str, Any]] = []

    grouped: dict[str, list[dict[str, Any]]] = {}
    pending_without_number: list[dict[str, Any]] = []
    for row in rows:
        source_rows.append(row)
        invoice_number = str(row.get("Numer faktury") or "").strip()
        if not invoice_number:
            pending_without_number.append(row)
            continue
        grouped.setdefault(invoice_number, []).append(row)

    for row in pending_without_number:
        issue = _row_issue("Numer faktury jest wymagany.", row, "Numer faktury")
        errors.append(issue)
        invalid_rows.append(_invalid_row(row, [issue], []))
        if mode is ImportMode.FAIL_FAST:
            raise FA3ImportError(issue)

    for invoice_number, invoice_rows in grouped.items():
        group_errors, group_warnings, draft = _draft_from_group(invoice_number, invoice_rows)
        errors.extend(group_errors)
        warnings.extend(group_warnings)
        if group_errors:
            for row in invoice_rows:
                invalid_rows.append(_invalid_row(row, group_errors, group_warnings))
            if mode is ImportMode.FAIL_FAST:
                raise FA3ImportError(group_errors[0])
        elif mode is not ImportMode.VALIDATE_ONLY and draft is not None:
            valid.append(draft)

    return FA3ImportResult(
        valid_drafts=valid,
        invalid_rows=invalid_rows,
        errors=errors,
        warnings=warnings,
        source_rows=source_rows,
        headers=headers,
    )


def _draft_from_group(
    invoice_number: str, rows: list[dict[str, Any]]
) -> tuple[list[FA3ValidationIssue], list[FA3ValidationIssue], FA3Draft | None]:
    errors: list[FA3ValidationIssue] = []
    warnings: list[FA3ValidationIssue] = []
    header = _inherited_header(rows)
    try:
        builder = FA3InvoiceBuilder(
            invoice_number=invoice_number,
            kind=FA3InvoiceKind.parse(_required(header, rows[0], "Typ faktury")),
            issue_date=_parse_date(
                _required(header, rows[0], "Data wystawienia"), "Data wystawienia"
            ),
            currency=str(_required(header, rows[0], "Waluta")).upper(),
            issue_place=str(header.get("Miejsce wystawienia") or ""),
            seller=FA3Party(
                name=str(_required(header, rows[0], "Nazwa sprzedawcy")),
                tax_id=str(_required(header, rows[0], "NIP sprzedawcy")),
                address=str(header.get("Adres sprzedawcy") or ""),
                country_code=str(header.get("Kraj sprzedawcy") or "PL"),
            ),
            buyer=FA3Party(
                name=str(_required(header, rows[0], "Nazwa nabywcy")),
                tax_id=str(_required(header, rows[0], "NIP nabywcy")),
                address=str(header.get("Adres nabywcy") or ""),
                country_code=str(header.get("Kraj nabywcy") or "PL"),
            ),
            payment_due_date=_parse_optional_date(
                header.get("Termin płatności"), "Termin płatności"
            ),
            payment_method=_optional_text(header.get("Forma płatności")),
            correction_reason=_optional_text(header.get("Przyczyna korekty")),
            corrected_invoice_number=_optional_text(header.get("Numer faktury korygowanej")),
            corrected_invoice_date=_parse_optional_date(
                header.get("Data faktury korygowanej"), "Data faktury korygowanej"
            ),
            corrected_ksef_number=_optional_text(header.get("Numer KSeF faktury korygowanej")),
            advance_invoice_number=_optional_text(header.get("Numer faktury zaliczkowej")),
            advance_ksef_number=_optional_text(header.get("Numer KSeF faktury zaliczkowej")),
            settlement_amount=_parse_optional_decimal(
                header.get("Kwota rozliczenia"), "Kwota rozliczenia"
            ),
        )
    except (ValueError, TypeError) as exc:
        errors.append(_row_issue(str(exc), rows[0]))
        return errors, warnings, None

    for row in rows:
        try:
            line = _line_from_row(row)
        except (ValueError, TypeError) as exc:
            errors.append(_row_issue(str(exc), row, _column_from_message(str(exc))))
            continue
        line_errors, line_warnings = line.validate()
        errors.extend(_locate_issue(issue, row) for issue in line_errors)
        warnings.extend(_locate_issue(issue, row) for issue in line_warnings)
        builder._lines.append(line)
        builder._warnings.extend(line_warnings)

    if errors:
        return errors, warnings, None
    try:
        draft = replace(builder.build(), warnings=tuple(warnings))
    except ValueError as exc:
        errors.append(_row_issue(str(exc), rows[0], _column_from_message(str(exc))))
        return errors, warnings, None
    return errors, warnings, draft


def _inherited_header(rows: list[dict[str, Any]]) -> dict[str, Any]:
    inherited: dict[str, Any] = {}
    for row in rows:
        for key, value in row.items():
            if key.startswith("__"):
                continue
            if value not in (None, "") and key not in inherited:
                inherited[key] = value
    return inherited


def _line_from_row(row: dict[str, Any]) -> FA3Line:
    return FA3Line(
        description=str(_required(row, row, "Opis pozycji")),
        quantity=decimal_from_value(_required(row, row, "Ilość"), field_name="Ilość"),
        unit=str(row.get("Jm") or "szt"),
        unit_net_price=decimal_from_value(
            _required(row, row, "Cena netto"), field_name="Cena netto"
        ),
        vat_rate=parse_vat_rate(_required(row, row, "VAT")),
        net_amount=_parse_optional_decimal(row.get("Wartość netto"), "Wartość netto"),
        vat_amount=_parse_optional_decimal(row.get("Kwota VAT"), "Kwota VAT"),
        gross_amount=_parse_optional_decimal(row.get("Wartość brutto"), "Wartość brutto"),
    )


def _required(source: dict[str, Any], row: dict[str, Any], column: str) -> Any:
    value = source.get(column)
    if value in (None, ""):
        cell = _cell(row, column)
        raise ValueError(f"{cell}: pole '{column}' jest wymagane.")
    return value


def _parse_date(value: Any, column: str) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value or "").strip()
    for fmt in ("%Y-%m-%d", "%d.%m.%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            pass
    raise ValueError(f"Pole '{column}' musi być datą w formacie RRRR-MM-DD.")


def _parse_optional_date(value: Any, column: str) -> date | None:
    if value in (None, ""):
        return None
    return _parse_date(value, column)


def _parse_optional_decimal(value: Any, column: str) -> Decimal | None:
    if value in (None, ""):
        return None
    return decimal_from_value(value, field_name=column)


def _optional_text(value: Any) -> str | None:
    text = str(value or "").strip()
    return text or None


def _row_issue(message: str, row: dict[str, Any], column: str | None = None) -> FA3ValidationIssue:
    return FA3ValidationIssue(
        message=message,
        invoice_number=_optional_text(row.get("Numer faktury")),
        row_number=int(row.get("__row_number__", 0)) or None,
        column=column,
        cell=_cell(row, column) if column else None,
    )


def _column_from_message(message: str) -> str | None:
    for column in REQUIRED_HEADERS:
        if f"'{column}'" in message or column in message:
            return column
    return None


def _locate_issue(issue: FA3ValidationIssue, row: dict[str, Any]) -> FA3ValidationIssue:
    column = issue.column
    return FA3ValidationIssue(
        message=issue.message,
        invoice_number=_optional_text(row.get("Numer faktury")),
        row_number=int(row.get("__row_number__", 0)) or None,
        column=column,
        cell=_cell(row, column) if column else None,
        severity=issue.severity,
    )


def _invalid_row(
    row: dict[str, Any],
    errors: list[FA3ValidationIssue],
    warnings: list[FA3ValidationIssue],
) -> FA3InvalidRow:
    return FA3InvalidRow(
        row_number=int(row.get("__row_number__", 0)),
        invoice_number=_optional_text(row.get("Numer faktury")),
        row_data={key: value for key, value in row.items() if not key.startswith("__")},
        errors=tuple(errors),
        warnings=tuple(warnings),
    )


def _cell(row: dict[str, Any], column: str | None) -> str:
    if not column:
        return ""
    try:
        from openpyxl.utils import get_column_letter

        index = list(REQUIRED_HEADERS).index(column) + 1
        return f"{get_column_letter(index)}{row.get('__row_number__')}"
    except (ImportError, ValueError):
        return f"wiersz {row.get('__row_number__')}, kolumna {column}"
